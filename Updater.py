import openpyxl
import datetime

def load_workbook(file_path):
    wb = openpyxl.load_workbook(file_path)
    return wb


def update_spreadsheet(wb, sheet_name, data, start_row=2):
    sheet = wb[sheet_name]
    
    for index, row_data in enumerate(data, start=start_row):
        for col_index, value in enumerate(row_data, start=1):
            sheet.cell(row=index, column=col_index).value = value
    
    print("Spreadsheet updated successfully!")


def save_workbook(wb, file_path):
    wb.save(file_path)
    print(f"Workbook saved at {file_path}")


def get_real_time_data():
   
    data = [
        [datetime.datetime.now(), "Product A", 150],
        [datetime.datetime.now(), "Product B", 200],
        [datetime.datetime.now(), "Product C", 300]
    ]
    return data


def main():
    file_path = "your_spreadsheet.xlsx"  
    sheet_name = "Sheet1"  
    
    wb = load_workbook(file_path)
    new_data = get_real_time_data()
    
    update_spreadsheet(wb, sheet_name, new_data)
    save_workbook(wb, file_path)

if __name__ == "__main__":
    main()

